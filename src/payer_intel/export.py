from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .schema import EXCEL_COLUMNS, PRODUCT_COLUMNS, PayerRecord


_VERDICT_STYLES: dict[str, tuple[str, str]] = {
    # verdict -> (fill_hex, font_hex)
    "Yes": ("C6EFCE", "276221"),
    "Likely": ("FFEB9C", "9C5700"),
    "No": ("FFC7CE", "9C0006"),
    "Unknown": ("F2F2F2", "808080"),
}


def _record_to_row(rec: PayerRecord) -> dict[str, str]:
    row = {
        "Payer Name": rec.payer_name,
        "Payer Type": rec.payer_type,
        "Source URLs": "\n".join(dict.fromkeys(rec.source_urls)),
        "Date Identified": rec.date_identified,
        "Confidence Score": rec.confidence.value,
        "BD Notes": rec.bd_notes,
        "Key Evidence": rec.key_evidence,
    }
    for col in PRODUCT_COLUMNS:
        row[col] = rec.verdicts.get(col, "Unknown")
    return row


def write_excel(records: Iterable[PayerRecord], out_dir: Path) -> Path:
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.utcnow().strftime("%Y%m%d")
    path = out_dir / f"Aarete_BD_Salesforce_Payer_Intelligence_{stamp}.xlsx"

    records_list = list(records)
    rows = [_record_to_row(r) for r in records_list]
    df = pd.DataFrame(rows, columns=EXCEL_COLUMNS)

    wb = Workbook()
    ws = wb.active
    ws.title = "Payer Intelligence"
    ws.append(EXCEL_COLUMNS)
    for r in df.itertuples(index=False):
        ws.append(list(r))

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for col_idx, _ in enumerate(EXCEL_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"

    widths = {
        "Payer Name": 28, "Payer Type": 14, "Source URLs": 55,
        "Date Identified": 16, "Confidence Score": 18, "BD Notes": 30,
        "Key Evidence": 60,
    }
    for col_idx, name in enumerate(EXCEL_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(name, 18)

    # Per-cell styling for product verdict columns + wrap_text on data rows
    product_col_indices = {col: EXCEL_COLUMNS.index(col) + 1 for col in PRODUCT_COLUMNS}
    wrap_align = Alignment(wrap_text=True, vertical="top")
    for row_num in range(2, len(rows) + 2):
        for col_idx in range(1, len(EXCEL_COLUMNS) + 1):
            ws.cell(row=row_num, column=col_idx).alignment = wrap_align
        for product, col_idx in product_col_indices.items():
            cell = ws.cell(row=row_num, column=col_idx)
            style = _VERDICT_STYLES.get(str(cell.value))
            if style:
                fill_hex, font_hex = style
                cell.fill = PatternFill("solid", fgColor=fill_hex)
                cell.font = Font(color=font_hex, bold=True)
                cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")

    # Conditional color scale on Confidence Score (kept for High/Medium/Low)
    if rows:
        conf_idx = EXCEL_COLUMNS.index("Confidence Score") + 1
        col_letter = get_column_letter(conf_idx)
        rng = f"{col_letter}2:{col_letter}{len(rows) + 1}"
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="equal", formula=['"High"'], fill=PatternFill("solid", fgColor="C6EFCE"))
        )
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="equal", formula=['"Medium"'], fill=PatternFill("solid", fgColor="FFEB9C"))
        )
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="equal", formula=['"Low"'], fill=PatternFill("solid", fgColor="FFC7CE"))
        )

    # Second sheet: Summary Dashboard — product × verdict counts
    summary = wb.create_sheet("Summary Dashboard")
    summary.append(["Salesforce Product", "Yes", "Likely", "No", "Unknown", "Total Payers"])
    for cell in summary[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    total_payers = len(records_list)
    for product in PRODUCT_COLUMNS:
        counts = {"Yes": 0, "Likely": 0, "No": 0, "Unknown": 0}
        for rec in records_list:
            verdict = rec.verdicts.get(product, "Unknown")
            counts[verdict if verdict in counts else "Unknown"] += 1
        summary.append([product, counts["Yes"], counts["Likely"], counts["No"], counts["Unknown"], total_payers])
    summary.column_dimensions["A"].width = 22
    for letter in ("B", "C", "D", "E", "F"):
        summary.column_dimensions[letter].width = 14
    summary.freeze_panes = "B2"
    # Color the Yes/Likely/Unknown columns lightly for readability
    for row_num in range(2, len(PRODUCT_COLUMNS) + 2):
        summary.cell(row=row_num, column=2).fill = PatternFill("solid", fgColor="E2EFDA")  # Yes
        summary.cell(row=row_num, column=3).fill = PatternFill("solid", fgColor="FFF2CC")  # Likely
        summary.cell(row=row_num, column=5).fill = PatternFill("solid", fgColor="F2F2F2")  # Unknown

    wb.save(path)
    return path
