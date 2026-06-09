import os
from pathlib import Path

import pytest
from openpyxl import load_workbook

from payer_intel.crew import run
from payer_intel.schema import EXCEL_COLUMNS


pytestmark = pytest.mark.skipif(
    os.getenv("RUN_LIVE_TESTS") != "1",
    reason="Live smoke test; set RUN_LIVE_TESTS=1 to enable (uses SearchApi.io + Bedrock).",
)


def test_smoke_full_pipeline(tmp_path: Path):
    root = Path(__file__).resolve().parents[1]
    seed = root / "data" / "seed_payers_smoke.csv"
    assert seed.exists(), f"missing seed: {seed}"

    out_path = run(seed, tmp_path)
    assert out_path.exists()

    wb = load_workbook(out_path)
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert header == EXCEL_COLUMNS

    # 3 payers in smoke seed + header row
    assert ws.max_row >= 4, f"expected ≥4 rows, got {ws.max_row}"

    # At least one cell across all product columns should be a positive verdict
    positives = {"Yes", "Likely"}
    product_cols = [
        EXCEL_COLUMNS.index(c) + 1
        for c in EXCEL_COLUMNS
        if c not in {"Payer Name", "Payer Type", "Source URLs", "Date Identified", "Confidence Score", "BD Notes"}
    ]
    found = False
    for row in ws.iter_rows(min_row=2, values_only=False):
        for idx in product_cols:
            if row[idx - 1].value in positives:
                found = True
                break
        if found:
            break
    assert found, "no positive Salesforce verdicts produced — investigate sourcing"
