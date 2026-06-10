from pathlib import Path

from openpyxl import load_workbook

from payer_intel.export import write_excel
from payer_intel.schema import EXCEL_COLUMNS, PRODUCT_COLUMNS, ConfidenceScore, PayerRecord


def test_excel_schema_and_freeze(tmp_path: Path):
    recs = [
        PayerRecord(
            payer_name="Humana Inc.",
            payer_type="National",
            domain="humana.com",
            verdicts={"Sales Cloud": "Yes", "Service Cloud": "Yes", "Health Cloud": "Likely"},
            source_urls=["https://a", "https://b"],
            date_identified="2025-10-01",
            confidence=ConfidenceScore.HIGH,
            bd_notes="recent case study + tech fingerprint",
            key_evidence="Humana shows multiple Salesforce footprints including a 2025 Health Cloud admin job posting and a Salesforce customer case study.",
        ),
        PayerRecord(payer_name="Centene Corporation", payer_type="National"),
    ]
    out = write_excel(recs, tmp_path)
    assert out.exists()
    assert out.name.startswith("Aarete_BD_Salesforce_Payer_Intelligence_")
    assert out.suffix == ".xlsx"

    wb = load_workbook(out)
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert header == EXCEL_COLUMNS, f"Header mismatch: {header}"
    assert "Key Evidence" in header
    assert ws.freeze_panes == "A2"

    # Row 2 — Humana — Sales Cloud cell == "Yes"
    sales_idx = EXCEL_COLUMNS.index("Sales Cloud") + 1
    assert ws.cell(row=2, column=sales_idx).value == "Yes"
    # Missing products default to "Unknown"
    data_cloud_idx = EXCEL_COLUMNS.index("Data Cloud") + 1
    assert ws.cell(row=2, column=data_cloud_idx).value == "Unknown"
    # Confidence in last 4 cols
    conf_idx = EXCEL_COLUMNS.index("Confidence Score") + 1
    assert ws.cell(row=2, column=conf_idx).value == "High"
    # Source URLs newline-joined
    src_idx = EXCEL_COLUMNS.index("Source URLs") + 1
    assert "\n" in ws.cell(row=2, column=src_idx).value
    # Key Evidence narrative populated
    ke_idx = EXCEL_COLUMNS.index("Key Evidence") + 1
    assert "Health Cloud" in (ws.cell(row=2, column=ke_idx).value or "")

    # Per-cell verdict styling: Sales Cloud (Yes) should have green fill
    yes_cell = ws.cell(row=2, column=sales_idx)
    assert (yes_cell.fill.fgColor.rgb or "").upper().endswith("C6EFCE")
    # Unknown cell should have grey fill
    unknown_cell = ws.cell(row=2, column=data_cloud_idx)
    assert (unknown_cell.fill.fgColor.rgb or "").upper().endswith("F2F2F2")


def test_summary_dashboard_sheet(tmp_path: Path):
    recs = [
        PayerRecord(
            payer_name="A",
            verdicts={"Sales Cloud": "Yes", "Service Cloud": "Likely"},
        ),
        PayerRecord(
            payer_name="B",
            verdicts={"Sales Cloud": "Yes"},
        ),
        PayerRecord(payer_name="C"),
    ]
    out = write_excel(recs, tmp_path)
    wb = load_workbook(out)
    assert "Summary Dashboard" in wb.sheetnames
    summary = wb["Summary Dashboard"]
    header = [c.value for c in summary[1]]
    assert header == ["Salesforce Product", "Yes", "Likely", "No", "Unknown", "Total Payers"]
    # One row per product
    product_rows = list(summary.iter_rows(min_row=2, values_only=True))
    assert len(product_rows) == len(PRODUCT_COLUMNS)
    by_product = {row[0]: row for row in product_rows}
    # Sales Cloud: 2 Yes, 0 Likely, 0 No, 1 Unknown, total 3
    assert by_product["Sales Cloud"][1:] == (2, 0, 0, 1, 3)
    # Service Cloud: 0 Yes, 1 Likely, 0 No, 2 Unknown, total 3
    assert by_product["Service Cloud"][1:] == (0, 1, 0, 2, 3)
